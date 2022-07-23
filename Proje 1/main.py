import PySimpleGUI as sg
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

while True:
    try:
        layout = [
            [sg.Text('How many students you want to enter for the Math class?')],
            [sg.Text('Number of students', size=(15, 1)), sg.InputText()],
            [sg.Submit()]
        ]

        window = sg.Window('INPUT DATA WINDOW', layout)
        event, values = window.read()
        window.close()
        if int(values[0]) <= 0:
            while True:
                sg.popup_error(
                    'INVALID INPUT!',
                    title='ERROR!',
                )
                layout = [
                    [sg.Text('Invalid input. How many students you want to enter for the Math class? Try again please. ')],
                    [sg.Text('Number of students', size=(15, 1)), sg.InputText()],
                    [sg.Submit()]
                ]

                window = sg.Window('INPUT DATA WINDOW', layout)
                event, values = window.read()
                window.close()
                break
        num_of_students = int(values[0])
        break

    except ValueError:
        sg.popup_error(
            'INVALID INPUT!',
            title='ERROR!',
        )

first_school_num = 2022001
school_num = []
student_name = []
student_surname = []
exam_result = []
passing_result = []
letter_result = []

sg.theme('Topanga')  # Add a touch of color


def pass_situation(result, name, surname):
    global window
    sg.theme('Topanga')  # Add a touch of color

    if result > 80:
        layout = [[sg.Text(f"{name} {surname} is passed the Math class with AA")], [sg.Button("OK")]]
        window = sg.Window("Passing Situation", layout, keep_on_top=True)
        passing_result.append("PASSED")
        letter_result.append("AA")

    elif result > 70:
        layout = [[sg.Text(f"{name} {surname} is passed the Math class with BB")], [sg.Button("OK")]]
        window = sg.Window("Passing Situation", layout, keep_on_top=True)
        passing_result.append("PASSED")
        letter_result.append("BB")

    elif result > 60:
        layout = [[sg.Text(f"{name} {surname} is passed the Math class with CC")], [sg.Button("OK")]]
        window = sg.Window("Passing Situation", layout, keep_on_top=True)
        passing_result.append("PASSED")
        letter_result.append("CC")

    elif result > 50:
        layout = [[sg.Text(f"{name} {surname} is passed the Math class with DD")], [sg.Button("OK")]]
        window = sg.Window("Passing Situation", layout, keep_on_top=True)
        passing_result.append("PASSED")
        letter_result.append("DD")

    elif result >= 0:
        layout = [[sg.Text(f"{name} {surname} is NOT passed the Math class!")], [sg.Button("OK")]]
        window = sg.Window("Passing Situation", layout, keep_on_top=True)
        passing_result.append("NOT PASSED")
        letter_result.append("FF")

    while True:
        event, values = window.read()
        if event == "OK" or event == sg.WIN_CLOSED:
            break

    window.close()


for i in range(num_of_students):
    school_num.append(first_school_num)
    first_school_num += 1
    name = input("Enter student's name : ")
    student_name.append(name)
    surname = input(f"Enter {name}'s surname : ")
    student_surname.append(surname)

    while True:
        try:
            result = int(input(f"Enter {name} {surname}'s exam result : "))
            break
        except ValueError:
            print("Invalid input. An integer value was expected. Try again.")

    if result < 0 or result > 100:
        while True:
            print("Invalid exam score. Try again please.")
            result = int(input(f"Enter {name} {surname}'s exam result : "))
            exam_result.append(result)
            break
    else:
        exam_result.append(result)
    pass_situation(result, name, surname)

data = {'Name': student_name,
        'Surname': student_surname,
        'School Number': school_num,
        'Result': exam_result,
        'Letter Result': letter_result,
        'Passing': passing_result
        }

df = pd.DataFrame(data)
print(df)

writer = pd.ExcelWriter('Math_Exam_Results.xlsx', engine='xlsxwriter')

df.to_excel(writer, sheet_name='Sheet1')

workbook = writer.book
worksheet = writer.sheets['Sheet1']
worksheet.insert_image('I4', '39928331.jpg')
worksheet.set_column_pixels(1, 6, 150)

fail_format = workbook.add_format({'bg_color': '#FFC7CE',
                                   'font_color': '#9C0006'})

pass_format = workbook.add_format({'bg_color': '#C6EFCE',
                                   'font_color': '#006100'})

start_col = 1
start_row = 1
end_col = df.shape[1] + 1
end_row = df.shape[0] + 1

worksheet.conditional_format(start_row, start_col, end_row, end_col,
                             {'type': 'text',
                              'criteria': 'containing',
                              'value': 'NOT PASSED',
                              'format': fail_format})

worksheet.conditional_format(start_row, start_col, end_row, end_col,
                             {'type': 'text',
                              'criteria': 'containing',
                              'value': 'PASSED',
                              'format': pass_format})

writer.save()

wb = load_workbook(filename="Math_Exam_Results.xlsx")

mysheet = wb['Sheet1']

logo = Image(r"39928331.jpg")
mysheet.add_image(logo, "H3")

for row in range(1, mysheet.max_row + 1):
    for col in range(1, mysheet.max_column + 1):
        cell = mysheet.cell(row, col)
        cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save("Math_Exam_Results.xlsx")
